"""
物资管理系统 - Flask Web Application
支持扫码入库/出库/盘点，商品二维码生成
"""

from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file, g
from flask_sqlalchemy import SQLAlchemy
from functools import wraps
import sqlite3
import uuid
import qrcode
import io
import base64
from datetime import datetime, timedelta
import os
from werkzeug.utils import secure_filename

UPLOAD_FOLDER = 'static/uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

app = Flask(__name__)
app.secret_key = os.urandom(24)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///inventory.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

@app.before_request
def inject_settings():
    """全局注入系统设置到所有模板"""
    g.system_name = SystemSettings.get('system_name', '物资管理系统')
    g.system_logo = SystemSettings.get('system_logo', '')
    g.login_background = SystemSettings.get('login_background', '')
    
    # 注入当前项目信息
    project_id = session.get('current_project_id')
    if project_id:
        project = Project.query.get(project_id)
        if project and project.is_active:
            g.current_project = project
        else:
            g.current_project = None
            session.pop('current_project_id', None)
    else:
        g.current_project = None
        # 自动选择第一个项目
        first_project = Project.query.filter_by(is_active=True).first()
        if first_project:
            session['current_project_id'] = first_project.id
            g.current_project = first_project

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def upload_image(file):
    if file and allowed_file(file.filename):
        ext = file.filename.rsplit('.', 1)[1].lower()
        filename = f"{uuid.uuid4().hex}.{ext}"
        file.save(os.path.join(UPLOAD_FOLDER, filename))
        return filename
    return None

db = SQLAlchemy(app)

# ============ 数据库模型 ============

class User(db.Model):
    """用户表"""
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password_hash = db.Column(db.String(128), nullable=False)
    role = db.Column(db.String(50), default='user')  # 角色名称
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Project(db.Model):
    """项目表"""
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    code = db.Column(db.String(50), unique=True, nullable=False)  # 项目代码
    description = db.Column(db.Text)
    logo = db.Column(db.String(200))  # 项目Logo
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Category(db.Model):
    """分类表"""
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'), nullable=True)  # 分类属于项目
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class SystemSettings(db.Model):
    """系统设置表"""
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(100), unique=True, nullable=False)
    value = db.Column(db.Text)

    @staticmethod
    def get(key, default=''):
        setting = SystemSettings.query.filter_by(key=key).first()
        return setting.value if setting else default
    
    @staticmethod
    def set(key, value):
        setting = SystemSettings.query.filter_by(key=key).first()
        if setting:
            setting.value = value
        else:
            setting = SystemSettings(key=key, value=value)
            db.session.add(setting)
        db.session.commit()

class Product(db.Model):
    """商品表"""
    id = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.Integer, db.ForeignKey('project.id'), nullable=False)  # 所属项目
    code = db.Column(db.String(50), unique=True, nullable=False)  # 商品编码（扫码用）
    name = db.Column(db.String(200), nullable=False)
    category_id = db.Column(db.Integer, db.ForeignKey('category.id'))
    unit = db.Column(db.String(20), default='件')  # 单位
    min_stock = db.Column(db.Integer, default=0)  # 最低库存预警
    current_stock = db.Column(db.Integer, default=0)
    location = db.Column(db.String(100))  # 存放位置
    image_url = db.Column(db.String(500))  # 商品图片路径
    remark = db.Column(db.Text)  # 备注
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    project = db.relationship('Project', backref='products')
    category = db.relationship('Category', backref='products')

class StockIn(db.Model):
    """入库记录表"""
    id = db.Column(db.Integer, primary_key=True)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False)
    quantity = db.Column(db.Integer, nullable=False)
    operator = db.Column(db.String(50), nullable=False)
    remark = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    product = db.relationship('Product', backref='stock_ins')

class StockOut(db.Model):
    """出库记录表"""
    id = db.Column(db.Integer, primary_key=True)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False)
    quantity = db.Column(db.Integer, nullable=False)
    operator = db.Column(db.String(50), nullable=False)
    remark =db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    product = db.relationship('Product', backref='stock_outs')

class StockCheck(db.Model):
    """盘点记录表"""
    id = db.Column(db.Integer, primary_key=True)
    product_id = db.Column(db.Integer, db.ForeignKey('product.id'), nullable=False)
    actual_stock = db.Column(db.Integer, nullable=False)  # 实际盘点数量
    system_stock = db.Column(db.Integer, nullable=False)  # 系统数量
    diff = db.Column(db.Integer, nullable=False)  # 差异
    operator = db.Column(db.String(50), nullable=False)
    remark = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    product = db.relationship('Product', backref='stock_checks')

# ============ 辅助函数 ============

def generate_code():
    """生成商品编码"""
    return 'P' + datetime.now().strftime('%Y%m%d%H%M%S') + str(uuid.uuid4().hex[:4]).upper()

def hash_password(pwd):
    """简单密码hash"""
    import hashlib
    return hashlib.sha256(pwd.encode()).hexdigest()

def verify_password(pwd, hashed):
    return hash_password(pwd) == hashed

def generate_qr_code(data):
    """生成二维码图片（返回base64）"""
    qr = qrcode.QRCode(version=1, box_size=10, border=4)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    
    buffer = io.BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    return base64.b64encode(buffer.getvalue()).decode()

def get_product_by_code(code):
    """通过编码查找商品"""
    return Product.query.filter_by(code=code).first()

# ============ 登录装饰器 ============

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# ============ 路由 ============

@app.route('/')
@login_required
def index():
    """仪表盘"""
    total_products = Product.query.count()
    low_stock_count = Product.query.filter(Product.current_stock <= Product.min_stock, Product.min_stock > 0).count()
    today = datetime.now().date()
    
    stock_in_today = StockIn.query.filter(
        db.func.date(StockIn.created_at) == today
    ).count()
    stock_out_today = StockOut.query.filter(
        db.func.date(StockOut.created_at) == today
    ).count()
    
    return render_template('index.html',
                         total_products=total_products,
                         low_stock_count=low_stock_count,
                         stock_in_today=stock_in_today,
                         stock_out_today=stock_out_today)

# ---- 登录 ----

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        user = User.query.filter_by(username=username).first()
        if user and verify_password(password, user.password_hash):
            session['user_id'] = user.id
            session['username'] = user.username
            session['role'] = user.role  # 保存角色到session
            return redirect(url_for('index'))
        return render_template('login.html', error='用户名或密码错误')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ---- 商品管理 ----

@app.route('/products')
@login_required
def products():
    """商品列表"""
    category_id = request.args.get('category_id', type=int)
    search = request.args.get('search', '')
    
    query = Product.query
    if category_id:
        query = query.filter_by(category_id=category_id)
    if search:
        query = query.filter(Product.name.contains(search) | Product.code.contains(search))
    
    products_list = query.order_by(Product.created_at.desc()).all()
    categories = Category.query.all()
    
    return render_template('products.html', 
                         products=products_list, 
                         categories=categories,
                         current_category=category_id,
                         search=search)

@app.route('/product/add', methods=['GET', 'POST'])
@login_required
def add_product():
    """添加商品"""
    if request.method == 'POST':
        name = request.form.get('name')
        category_id = request.form.get('category_id', type=int)
        unit = request.form.get('unit', '件')
        min_stock = request.form.get('min_stock', type=int, default=0)
        location = request.form.get('location', '')
        remark = request.form.get('remark', '')
        
        # 处理图片上传
        image_url = None
        if 'image' in request.files:
            file = request.files['image']
            image_url = upload_image(file)
        
        product = Product(
            code=generate_code(),
            name=name,
            category_id=category_id,
            unit=unit,
            min_stock=min_stock,
            location=location,
            image_url=image_url,
            remark=remark
        )
        db.session.add(product)
        db.session.commit()
        return redirect(url_for('products'))
    
    categories = Category.query.all()
    return render_template('product_form.html', product=None, categories=categories)

@app.route('/product/<int:product_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_product(product_id):
    """编辑商品"""
    product = Product.query.get_or_404(product_id)
    
    if request.method == 'POST':
        product.name = request.form.get('name')
        product.category_id = request.form.get('category_id', type=int)
        product.unit = request.form.get('unit', '件')
        product.min_stock = request.form.get('min_stock', type=int, default=0)
        product.location = request.form.get('location', '')
        product.remark = request.form.get('remark', '')
        
        # 处理图片上传（可选）
        if 'image' in request.files:
            file = request.files['image']
            if file and file.filename:
                new_image = upload_image(file)
                if new_image:
                    product.image_url = new_image
        
        db.session.commit()
        return redirect(url_for('products'))
    
    categories = Category.query.all()
    return render_template('product_form.html', product=product, categories=categories)

@app.route('/product/<int:product_id>/delete', methods=['POST'])
@login_required
def delete_product(product_id):
    """删除商品"""
    product = Product.query.get_or_404(product_id)
    db.session.delete(product)
    db.session.commit()
    return redirect(url_for('products'))

@app.route('/product/<int:product_id>/qrcode')
@login_required
def product_qrcode(product_id):
    """生成商品二维码"""
    product = Product.query.get_or_404(product_id)
    qr_data = f"INV:{product.code}"
    qr_base64 = generate_qr_code(qr_data)
    return render_template('qrcode.html', product=product, qr_base64=qr_base64)

# ---- 入库 ----

@app.route('/stock-in')
@login_required
def stock_in_page():
    """入库页面"""
    products = Product.query.order_by(Product.name).all()
    return render_template('stock_in.html', products=products)

@app.route('/stock-in/import', methods=['POST'])
@login_required
def import_stock_in():
    """批量导入入库"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '没有上传文件'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '没有选择文件'})
    
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(file)
        ws = wb.active
        
        # 跳过表头，从第二行开始读取
        success_count = 0
        error_list = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            code = str(row[0]).strip() if row[0] else ''
            quantity = row[1]
            remark = str(row[2]).strip() if row[2] else ''
            
            if not code:
                continue
            
            # 支持扫码格式
            if code.startswith('INV:'):
                code = code[4:]
            
            product = get_product_by_code(code)
            if not product:
                error_list.append(f'第{row_idx}行: 商品编码不存在 {code}')
                continue
            
            if quantity is None or quantity <= 0:
                error_list.append(f'第{row_idx}行: 数量无效 {quantity}')
                continue
            
            stock_in = StockIn(
                product_id=product.id,
                quantity=int(quantity),
                operator=session['username'],
                remark=remark
            )
            product.current_stock += int(quantity)
            db.session.add(stock_in)
            success_count += 1
        
        db.session.commit()
        
        message = f'成功导入 {success_count} 条记录'
        if error_list:
            message += f'，{len(error_list)} 条失败'
        
        return jsonify({
            'success': success_count > 0,
            'message': message,
            'errors': error_list[:10]  # 最多返回10条错误
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'导入失败: {str(e)}'})


@app.route('/stock-in/import-jd', methods=['POST'])
def import_jd_order():
    """批量导入京东工采订单"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '没有上传文件'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '没有选择文件'})
    
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(file)
        ws = wb.active
        
        # 获取表头
        headers = [str(h).strip() if h else '' for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        
        # 查找关键列索引
        col_mapping = {}
        for idx, header in enumerate(headers):
            if header in ['商品编码', '物料编码']:
                col_mapping['code'] = idx
            elif header == '物料名称':
                col_mapping['material_name'] = idx
            elif header == '商品名称':
                col_mapping['product_name'] = idx
            elif header == '数量':
                col_mapping['quantity'] = idx
        
        if 'code' not in col_mapping and 'material_name' not in col_mapping and 'product_name' not in col_mapping:
            return jsonify({'success': False, 'message': '无法识别京东订单格式，请确保包含商品编码或商品名称列'})
        
        success_count = 0
        error_list = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # 尝试获取商品编码
            code = ''
            if 'code' in col_mapping:
                code = str(row[col_mapping['code']]).strip() if row[col_mapping['code']] else ''
            
            # 如果没有编码，尝试用商品名称匹配
            product = None
            if code:
                if code.startswith('INV:'):
                    code = code[4:]
                product = get_product_by_code(code)
            
            # 如果没找到，尝试用物料名称匹配
            if not product and 'material_name' in col_mapping:
                material_name = str(row[col_mapping['material_name']]).strip() if row[col_mapping['material_name']] else ''
                if material_name:
                    product = Product.query.filter_by(name=material_name).first()
            
            # 再尝试用商品名称匹配
            if not product and 'product_name' in col_mapping:
                product_name = str(row[col_mapping['product_name']]).strip() if row[col_mapping['product_name']] else ''
                if product_name:
                    product = Product.query.filter_by(name=product_name).first()
            
            if not product:
                name_hint = ''
                if 'material_name' in col_mapping:
                    name_hint = str(row[col_mapping['material_name']]).strip()[:20]
                elif 'product_name' in col_mapping:
                    name_hint = str(row[col_mapping['product_name']]).strip()[:20]
                error_list.append(f'第{row_idx}行: 商品未找到 {name_hint}')
                continue
            
            # 获取数量
            quantity = row[col_mapping['quantity']] if 'quantity' in col_mapping else None
            if quantity is None or quantity <= 0:
                error_list.append(f'第{row_idx}行: 数量无效')
                continue
            
            # 创建入库记录
            stock_in = StockIn(
                product_id=product.id,
                quantity=int(quantity),
                operator=session['username'],
                remark='京东工采订单导入'
            )
            product.current_stock += int(quantity)
            db.session.add(stock_in)
            success_count += 1
        
        db.session.commit()
        
        message = f'成功导入 {success_count} 条记录'
        if error_list:
            message += f'，{len(error_list)} 条失败'
        
        return jsonify({
            'success': success_count > 0,
            'message': message,
            'errors': error_list[:10]
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'导入失败: {str(e)}'})


@app.route('/stock-in/do', methods=['POST'])
@login_required
def do_stock_in():
    """执行入库"""
    code = request.form.get('code', '').strip()
    quantity = request.form.get('quantity', type=int)
    remark = request.form.get('remark', '')
    
    # 支持扫码（code格式: INV:xxxxx）
    if code.startswith('INV:'):
        code = code[4:]
    
    product = get_product_by_code(code)
    if not product:
        return jsonify({'success': False, 'message': f'商品编码不存在: {code}'})
    
    if quantity <= 0:
        return jsonify({'success': False, 'message': '数量必须大于0'})
    
    # 入库记录
    stock_in = StockIn(
        product_id=product.id,
        quantity=quantity,
        operator=session['username'],
        remark=remark
    )
    product.current_stock += quantity
    
    db.session.add(stock_in)
    db.session.commit()
    
    return jsonify({
        'success': True, 
        'message': f'入库成功: {product.name} x {quantity}',
        'product_name': product.name,
        'new_stock': product.current_stock
    })

@app.route('/stock-in/records')
@login_required
def stock_in_records():
    """入库记录"""
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    
    query = StockIn.query.order_by(StockIn.created_at.desc())
    
    if date_from:
        query = query.filter(db.func.date(StockIn.created_at) >= date_from)
    if date_to:
        query = query.filter(db.func.date(StockIn.created_at) <= date_to)
    
    records = query.limit(100).all()
    return render_template('stock_records.html', records=records, type='in')

# ---- 出库 ----

@app.route('/stock-out')
@login_required
def stock_out_page():
    """出库页面"""
    products = Product.query.order_by(Product.name).all()
    categories = Category.query.order_by(Category.name).all()
    return render_template('stock_out.html', products=products, categories=categories)

@app.route('/stock-out/import', methods=['POST'])
@login_required
def import_stock_out():
    """批量导入出库"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '没有上传文件'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '没有选择文件'})
    
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(file)
        ws = wb.active
        
        success_count = 0
        error_list = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            code = str(row[0]).strip() if row[0] else ''
            quantity = row[1]
            remark = str(row[2]).strip() if row[2] else ''
            
            if not code:
                continue
            
            if code.startswith('INV:'):
                code = code[4:]
            
            product = get_product_by_code(code)
            if not product:
                error_list.append(f'第{row_idx}行: 商品编码不存在 {code}')
                continue
            
            if quantity is None or quantity <= 0:
                error_list.append(f'第{row_idx}行: 数量无效 {quantity}')
                continue
            
            if product.current_stock < quantity:
                error_list.append(f'第{row_idx}行: 库存不足 (当前 {product.current_stock})')
                continue
            
            stock_out = StockOut(
                product_id=product.id,
                quantity=int(quantity),
                operator=session['username'],
                remark=remark
            )
            product.current_stock -= int(quantity)
            db.session.add(stock_out)
            success_count += 1
        
        db.session.commit()
        
        message = f'成功导出 {success_count} 条记录'
        if error_list:
            message += f'，{len(error_list)} 条失败'
        
        return jsonify({
            'success': success_count > 0,
            'message': message,
            'errors': error_list[:10]
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'导入失败: {str(e)}'})

@app.route('/stock-out/do', methods=['POST'])
@login_required
def do_stock_out():
    """执行出库"""
    code = request.form.get('code', '').strip()
    quantity = request.form.get('quantity', type=int)
    remark = request.form.get('remark', '')
    
    if code.startswith('INV:'):
        code = code[4:]
    
    product = get_product_by_code(code)
    if not product:
        return jsonify({'success': False, 'message': f'商品编码不存在: {code}'})
    
    if quantity <= 0:
        return jsonify({'success': False, 'message': '数量必须大于0'})
    
    if product.current_stock < quantity:
        return jsonify({'success': False, 'message': f'库存不足，当前库存: {product.current_stock}'})
    
    stock_out = StockOut(
        product_id=product.id,
        quantity=quantity,
        operator=session['username'],
        remark=remark
    )
    product.current_stock -= quantity
    
    db.session.add(stock_out)
    db.session.commit()
    
    return jsonify({
        'success': True,
        'message': f'出库成功: {product.name} x {quantity}',
        'product_name': product.name,
        'new_stock': product.current_stock
    })

@app.route('/stock-out/records')
@login_required
def stock_out_records():
    """出库记录"""
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    
    query = StockOut.query.order_by(StockOut.created_at.desc())
    
    if date_from:
        query = query.filter(db.func.date(StockOut.created_at) >= date_from)
    if date_to:
        query = query.filter(db.func.date(StockOut.created_at) <= date_to)
    
    records = query.limit(100).all()
    return render_template('stock_records.html', records=records, type='out')

# ---- 盘点 ----

@app.route('/stock-check')
@login_required
def stock_check_page():
    """盘点页面"""
    products = Product.query.order_by(Product.name).all()
    categories = Category.query.order_by(Category.name).all()
    return render_template('stock_check.html', products=products, categories=categories)

@app.route('/stock-check/do', methods=['POST'])
@login_required
def do_stock_check():
    """执行盘点"""
    code = request.form.get('code', '').strip()
    actual_stock = request.form.get('actual_stock', type=int)
    remark = request.form.get('remark', '')
    
    if code.startswith('INV:'):
        code = code[4:]
    
    product = get_product_by_code(code)
    if not product:
        return jsonify({'success': False, 'message': f'商品编码不存在: {code}'})
    
    if actual_stock < 0:
        return jsonify({'success': False, 'message': '数量不能为负'})
    
    diff = actual_stock - product.current_stock
    
    stock_check = StockCheck(
        product_id=product.id,
        actual_stock=actual_stock,
        system_stock=product.current_stock,
        diff=diff,
        operator=session['username'],
        remark=remark
    )
    
    product.current_stock = actual_stock
    
    db.session.add(stock_check)
    db.session.commit()
    
    diff_text = '正常' if diff == 0 else (f'盘盈 +{diff}' if diff > 0 else f'盘亏 {diff}')
    
    return jsonify({
        'success': True,
        'message': f'盘点完成: {product.name}',
        'diff_text': diff_text,
        'new_stock': product.current_stock
    })

@app.route('/stock-check/records')
@login_required
def stock_check_records():
    """盘点记录"""
    records = StockCheck.query.order_by(StockCheck.created_at.desc()).limit(100).all()
    return render_template('stock_check_records.html', records=records)

# ---- 分类管理 ----

@app.route('/categories')
@login_required
def categories():
    """分类列表"""
    categories_list = Category.query.order_by(Category.name).all()
    return render_template('categories.html', categories=categories_list)

@app.route('/category/add', methods=['POST'])
@login_required
def add_category():
    """添加分类"""
    name = request.form.get('name')
    category = Category(name=name)
    db.session.add(category)
    db.session.commit()
    return redirect(url_for('categories'))

@app.route('/category/<int:category_id>/delete', methods=['POST'])
@login_required
def delete_category(category_id):
    """删除分类"""
    category = Category.query.get_or_404(category_id)
    # 检查是否有商品使用此分类
    if Product.query.filter_by(category_id=category_id).first():
        return jsonify({'success': False, 'message': '该分类下有商品，无法删除'})
    db.session.delete(category)
    db.session.commit()
    return redirect(url_for('categories'))

# ---- 预警 ----

@app.route('/alerts')
@login_required
def alerts():
    """库存预警（低库存）"""
    low_stock = Product.query.filter(
        Product.current_stock <= Product.min_stock,
        Product.min_stock > 0
    ).order_by(Product.current_stock).all()
    return render_template('alerts.html', products=low_stock, alert_type='low')

@app.route('/high-stock-alerts')
@login_required
def high_stock_alerts():
    """高库存预警"""
    # 高库存 = 库存超过最高库存（暂时用 min_stock * 3 作为默认值，或者需要加 max_stock 字段）
    # 这里先简单实现：如果设置了预警值且库存超过预警值的 3 倍
    high_stock = Product.query.filter(
        Product.min_stock > 0,
        Product.current_stock >= Product.min_stock * 3
    ).order_by(Product.current_stock.desc()).all()
    return render_template('alerts.html', products=high_stock, alert_type='high')

# ---- 系统设置 ----

@app.route('/settings')
@login_required
def settings():
    """系统设置"""
    if session.get('role') != 'admin':
        return '无权限', 403
    
    # 获取当前设置
    system_name = SystemSettings.get('system_name', '物资管理系统')
    system_logo = SystemSettings.get('system_logo', '')
    login_background = SystemSettings.get('login_background', '')
    
    return render_template('settings.html', 
                         system_name=system_name,
                         system_logo=system_logo,
                         login_background=login_background)

@app.route('/settings/upload-logo', methods=['POST'])
@login_required
def upload_logo():
    """上传系统Logo"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'message': '无权限'})
    
    if 'logo' not in request.files:
        return jsonify({'success': False, 'message': '没有文件'})
    
    file = request.files['logo']
    if file.filename == '':
        return jsonify({'success': False, 'message': '没有选择文件'})
    
    if file and allowed_file(file.filename):
        ext = file.filename.rsplit('.', 1)[1].lower()
        filename = f"logo.{ext}"
        file.save(os.path.join(UPLOAD_FOLDER, filename))
        SystemSettings.set('system_logo', filename)
        return jsonify({'success': True, 'message': 'Logo上传成功'})
    
    return jsonify({'success': False, 'message': '不支持的文件格式'})

@app.route('/settings/upload-background', methods=['POST'])
@login_required
def upload_background():
    """上传登录页背景"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'message': '无权限'})
    
    if 'background' not in request.files:
        return jsonify({'success': False, 'message': '没有文件'})
    
    file = request.files['background']
    if file.filename == '':
        return jsonify({'success': False, 'message': '没有选择文件'})
    
    if file and allowed_file(file.filename):
        ext = file.filename.rsplit('.', 1)[1].lower()
        filename = f"background.{ext}"
        file.save(os.path.join(UPLOAD_FOLDER, filename))
        SystemSettings.set('login_background', filename)
        return jsonify({'success': True, 'message': '背景上传成功'})
    
    return jsonify({'success': False, 'message': '不支持的文件格式'})

@app.route('/settings/update', methods=['POST'])
@login_required
def update_settings():
    """更新系统设置"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'message': '无权限'})
    
    system_name = request.form.get('system_name', '')
    
    SystemSettings.set('system_name', system_name)
    
    return jsonify({'success': True, 'message': '设置已保存'})

@app.route('/settings/reset-logo', methods=['POST'])
@login_required
def reset_logo():
    """删除Logo"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'message': '无权限'})
    SystemSettings.set('system_logo', '')
    return jsonify({'success': True, 'message': 'Logo已删除'})

@app.route('/settings/reset-background', methods=['POST'])
@login_required
def reset_background():
    """删除登录背景"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'message': '无权限'})
    SystemSettings.set('login_background', '')
    return jsonify({'success': True, 'message': '背景已删除'})

# ---- 权限管理 ----

# 默认权限配置
# 默认权限配置
DEFAULT_PERMISSIONS = {
    'stock_in': True,
    'stock_out': True,
    'stock_check': True,
    'products_view': True,
    'products_edit': False,
    'products_delete': False,
    'categories_edit': False,
    'users_view': False,
    'users_edit': False,
    'settings_view': False,
    'settings_edit': False,
    'reports_view': True,
}

PERMISSION_LABELS = {
    'stock_in': '入库操作',
    'stock_out': '出库操作',
    'stock_check': '盘点操作',
    'products_view': '查看商品',
    'products_edit': '编辑商品',
    'products_delete': '删除商品',
    'categories_edit': '管理分类',
    'users_view': '查看用户',
    'users_edit': '管理用户',
    'settings_view': '查看设置',
    'settings_edit': '修改设置',
    'reports_view': '查看报表',
}

PERMISSION_GROUPS = {
    '库存操作': ['stock_in', 'stock_out', 'stock_check'],
    '商品管理': ['products_view', 'products_edit', 'products_delete'],
    '系统管理': ['categories_edit', 'users_view', 'users_edit', 'settings_view', 'settings_edit'],
    '报表': ['reports_view'],
}

def get_roles():
    """获取所有角色"""
    roles_data = SystemSettings.get('roles', '')
    if roles_data:
        import json
        return json.loads(roles_data)
    # 默认角色
    return [
        {'name': 'admin', 'is_admin': True},
        {'name': 'user', 'is_admin': False}
    ]

def save_roles(roles):
    """保存角色列表"""
    import json
    SystemSettings.set('roles', json.dumps(roles))

def get_permissions(role_name):
    """获取角色的权限"""
    all_perms = SystemSettings.get('permissions', '')
    if all_perms:
        import json
        perms = json.loads(all_perms)
        return perms.get(role_name, DEFAULT_PERMISSIONS.copy())
    return DEFAULT_PERMISSIONS.copy()

def save_permissions_for_role(role_name, perms):
    """保存角色的权限"""
    import json
    all_perms = SystemSettings.get('permissions', '')
    if all_perms:
        perms_dict = json.loads(all_perms)
    else:
        perms_dict = {}
    perms_dict[role_name] = perms
    SystemSettings.set('permissions', json.dumps(perms_dict))

def check_permission(perm_key):
    """检查当前用户是否有某权限"""
    import json
    role = session.get('role', 'user')
    perms = get_permissions(role)
    
    # 管理员拥有所有权限
    roles = get_roles()
    for r in roles:
        if r['name'] == role and r.get('is_admin', False):
            return True
    
    return perms.get(perm_key, False)

@app.route('/roles')
@login_required
def roles():
    """角色管理"""
    if session.get('role') != 'admin':
        return '无权限', 403
    roles_list = get_roles()
    # 计算每个角色的用户数
    users = User.query.all()
    for role in roles_list:
        role['user_count'] = sum(1 for u in users if u.role == role['name'])
    return render_template('roles.html', roles=roles_list)



# ============ 项目管理 ============
@app.route('/projects')
@login_required
def projects():
    """项目管理"""
    if session.get('role') != 'admin':
        return '无权限', 403
    
    # 获取用户管理的项目
    role = session.get('role', 'user')
    roles = get_roles()
    is_admin = any(r['name'] == role and r.get('is_admin', False) for r in roles)
    
    if is_admin:
        projects_list = Project.query.order_by(Project.created_at.desc()).all()
    else:
        # 普通用户只看有权限的项目
        from sqlalchemy import text
        user_project_ids = db.session.execute(
            text("SELECT project_id FROM user_project_roles WHERE username = :username"),
            {"username": session.get('username')}
        ).fetchall()
        projects_list = Project.query.filter(
            Project.id.in_([r[0] for r in user_project_ids]),
            Project.is_active == True
        ).all() if user_project_ids else []
    
    return render_template('projects.html', projects=projects_list)

@app.route('/project/add', methods=['POST'])
@login_required
def add_project():
    """添加项目"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'message': '无权限'})
    
    name = request.form.get('name', '').strip()
    code = request.form.get('code', '').strip().upper()
    description = request.form.get('description', '').strip()
    
    if not name or not code:
        return jsonify({'success': False, 'message': '项目名称和代码不能为空'})
    
    # 检查代码唯一性
    if Project.query.filter_by(code=code).first():
        return jsonify({'success': False, 'message': '项目代码已存在'})
    
    logo = None
    if 'logo' in request.files:
        file = request.files['logo']
        if file and file.filename and allowed_file(file.filename):
            logo = upload_image(file)
    
    project = Project(name=name, code=code, description=description, logo=logo)
    db.session.add(project)
    db.session.commit()
    
    return jsonify({'success': True, 'message': f'项目 "{name}" 创建成功'})

@app.route('/project/<int:project_id>/edit', methods=['POST'])
@login_required
def edit_project(project_id):
    """编辑项目"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'message': '无权限'})
    
    project = Project.query.get_or_404(project_id)
    
    name = request.form.get('name', '').strip()
    code = request.form.get('code', '').strip().upper()
    description = request.form.get('description', '').strip()
    is_active = request.form.get('is_active') == 'on'
    
    # 检查代码唯一性（排除自己）
    if Project.query.filter(Project.code == code, Project.id != project_id).first():
        return jsonify({'success': False, 'message': '项目代码已存在'})
    
    project.name = name
    project.code = code
    project.description = description
    project.is_active = is_active
    
    # 处理logo上传
    if 'logo' in request.files:
        file = request.files['logo']
        if file and file.filename and allowed_file(file.filename):
            project.logo = upload_image(file)
    
    db.session.commit()
    
    return jsonify({'success': True, 'message': '项目已更新'})

@app.route('/project/<int:project_id>/delete', methods=['POST'])
@login_required
def delete_project(project_id):
    """删除项目"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'message': '无权限'})
    
    project = Project.query.get_or_404(project_id)
    
    # 检查是否有商品
    if Product.query.filter_by(project_id=project_id).first():
        return jsonify({'success': False, 'message': '该项目下有商品，请先转移或删除'})
    
    db.session.delete(project)
    db.session.commit()
    
    return jsonify({'success': True, 'message': '项目已删除'})

@app.route('/project/<int:project_id>/users')
@login_required
def project_users(project_id):
    """项目用户管理"""
    if session.get('role') != 'admin':
        return '无权限', 403
    
    project = Project.query.get_or_404(project_id)
    users = User.query.all()
    roles = get_roles()
    
    # 获取项目用户和角色
    from sqlalchemy import text
    user_roles = db.session.execute(
        text("SELECT username, role FROM user_project_roles WHERE project_id = :project_id"),
        {"project_id": project_id}
    ).fetchall()
    user_role_dict = {ur[0]: ur[1] for ur in user_roles}
    
    return render_template('project_users.html', project=project, users=users, roles=roles, user_role_dict=user_role_dict)

@app.route('/project/<int:project_id>/user/add', methods=['POST'])
@login_required
def add_project_user(project_id):
    """添加项目用户"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'message': '无权限'})
    
    username = request.form.get('username', '').strip()
    role = request.form.get('role', 'viewer').strip()
    
    if not username:
        return jsonify({'success': False, 'message': '请选择用户'})
    
    # 检查用户是否存在
    user = User.query.filter_by(username=username).first()
    if not user:
        return jsonify({'success': False, 'message': '用户不存在'})
    
    from sqlalchemy import text
    # 检查是否已添加
    existing = db.session.execute(
        text("SELECT 1 FROM user_project_roles WHERE project_id = :project_id AND username = :username"),
        {"project_id": project_id, "username": username}
    ).fetchone()
    
    if existing:
        # 更新角色
        db.session.execute(
            text("UPDATE user_project_roles SET role = :role WHERE project_id = :project_id AND username = :username"),
            {"project_id": project_id, "username": username, "role": role}
        )
    else:
        db.session.execute(
            text("INSERT INTO user_project_roles (project_id, username, role) VALUES (:project_id, :username, :role)"),
            {"project_id": project_id, "username": username, "role": role}
        )
    
    db.session.commit()
    return jsonify({'success': True, 'message': f'已添加用户 "{username}" 到项目'})

@app.route('/project/<int:project_id>/user/<username>/remove', methods=['POST'])
@login_required
def remove_project_user(project_id, username):
    """移除项目用户"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'message': '无权限'})
    
    from sqlalchemy import text
    db.session.execute(
        text("DELETE FROM user_project_roles WHERE project_id = :project_id AND username = :username"),
        {"project_id": project_id, "username": username}
    )
    db.session.commit()
    
    return jsonify({'success': True, 'message': f'已移除用户 "{username}"'})

@app.route('/api/current-project', methods=['GET'])
@login_required
def get_current_project():
    """获取当前项目"""
    project_id = session.get('current_project_id')
    if project_id:
        project = Project.query.get(project_id)
        if project and project.is_active:
            return jsonify({'success': True, 'project_id': project_id, 'project_name': project.name, 'project_code': project.code})
    
    # 如果没有或无效，返回第一个有权限的项目
    username = session.get('username')
    from sqlalchemy import text
    
    # 管理员看所有项目
    role = session.get('role')
    roles_data = get_roles()
    is_admin = any(r['name'] == role and r.get('is_admin', False) for r in roles_data)
    
    if is_admin:
        project = Project.query.filter_by(is_active=True).first()
    else:
        result = db.session.execute(
            text("SELECT p.id, p.name, p.code FROM projects p JOIN user_project_roles upr ON p.id = upr.project_id WHERE upr.username = :username AND p.is_active = 1 LIMIT 1"),
            {"username": username}
        ).fetchone()
        project = Project.query.get(result[0]) if result else None
    
    if project:
        session['current_project_id'] = project.id
        return jsonify({'success': True, 'project_id': project.id, 'project_name': project.name, 'project_code': project.code})
    
    return jsonify({'success': False, 'message': '没有可用的项目'})

@app.route('/api/projects', methods=['GET'])
@login_required
def list_projects():
    """获取所有可访问项目"""
    username = session.get('username')
    role = session.get('role')
    
    roles_data = get_roles()
    is_admin = any(r['name'] == role and r.get('is_admin', False) for r in roles_data)
    
    if is_admin:
        projects = Project.query.filter_by(is_active=True).all()
    else:
        from sqlalchemy import text
        result = db.session.execute(
            text("SELECT p.id, p.name, p.code FROM projects p JOIN user_project_roles upr ON p.id = upr.project_id WHERE upr.username = :username AND p.is_active = 1"),
            {"username": username}
        ).fetchall()
        projects = [Project(id=r[0], name=r[1], code=r[2]) for r in result]
    
    return jsonify({
        'success': True,
        'projects': [{'id': p.id, 'name': p.name, 'code': p.code} for p in projects]
    })

@app.route('/api/switch-project/<int:project_id>', methods=['POST'])
@login_required
def switch_project(project_id):
    """切换当前项目"""
    username = session.get('username')
    role = session.get('role')
    
    # 检查权限
    roles_data = get_roles()
    is_admin = any(r['name'] == role and r.get('is_admin', False) for r in roles_data)
    
    if not is_admin:
        from sqlalchemy import text
        has_access = db.session.execute(
            text("SELECT 1 FROM user_project_roles WHERE project_id = :project_id AND username = :username"),
            {"project_id": project_id, "username": username}
        ).fetchone()
        if not has_access:
            return jsonify({'success': False, 'message': '无权访问此项目'})
    
    project = Project.query.get(project_id)
    if not project or not project.is_active:
        return jsonify({'success': False, 'message': '项目不存在或已停用'})
    
    session['current_project_id'] = project_id
    return jsonify({'success': True, 'message': f'已切换到项目 "{project.name}"'})

@app.template_filter('get_user_project_role')
def get_user_project_role(project_id, username):
    """获取用户在项目中的角色"""
    from sqlalchemy import text
    result = db.session.execute(
        text("SELECT role FROM user_project_roles WHERE project_id = :project_id AND username = :username"),
        {"project_id": project_id, "username": username}
    ).fetchone()
    return result[0] if result else None


@app.route('/role/add', methods=['POST'])
@login_required
def add_role():
    """添加角色"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'message': '无权限'})
    
    name = request.form.get('name', '').strip()
    if not name:
        return jsonify({'success': False, 'message': '角色名称不能为空'})
    
    roles = get_roles()
    if any(r['name'] == name for r in roles):
        return jsonify({'success': False, 'message': '角色已存在'})
    
    roles.append({'name': name, 'is_admin': False})
    save_roles(roles)
    # 新角色默认使用默认权限
    save_permissions_for_role(name, DEFAULT_PERMISSIONS.copy())
    
    return jsonify({'success': True, 'message': '角色添加成功'})

@app.route('/role/<role_name>/delete', methods=['POST'])
@login_required
def delete_role(role_name):
    """删除角色"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'message': '无权限'})
    
    roles = get_roles()
    # 不能删除 admin 和 user
    if role_name in ['admin', 'user']:
        return jsonify({'success': False, 'message': '不能删除内置角色'})
    
    # 检查是否有用户使用此角色
    if User.query.filter_by(role=role_name).first():
        return jsonify({'success': False, 'message': '有用户使用此角色，无法删除'})
    
    roles = [r for r in roles if r['name'] != role_name]
    save_roles(roles)
    
    return jsonify({'success': True, 'message': '角色已删除'})

@app.route('/permissions')
@login_required
def permissions():
    """权限管理"""
    if session.get('role') != 'admin':
        return '无权限', 403
    
    roles_list = get_roles()
    permissions_data = {}
    for role in roles_list:
        permissions_data[role['name']] = get_permissions(role['name'])
    
    return render_template('permissions.html', 
                         roles=roles_list,
                         permissions=permissions_data,
                         permission_labels=PERMISSION_LABELS,
                         permission_groups=PERMISSION_GROUPS)

@app.route('/permissions/save', methods=['POST'])
@login_required
def save_permissions():
    """保存权限配置"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'message': '无权限'})
    
    import json
    roles = get_roles()
    
    for role in roles:
        role_name = role['name']
        perms = {}
        for perm_key in PERMISSION_LABELS.keys():
            perms[perm_key] = request.form.get(f'{role_name}_{perm_key}') == 'on'
        save_permissions_for_role(role_name, perms)
    
    return jsonify({'success': True, 'message': '权限配置已保存'})

# ---- 用户管理 ----

@app.route('/users')
@login_required
def users():
    """用户列表（仅管理员）"""
    if session.get('role') != 'admin':
        return '无权限', 403
    users_list = User.query.all()
    return render_template('users.html', users=users_list)

@app.route('/user/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    """修改自己的密码"""
    if request.method == 'POST':
        old_password = request.form.get('old_password')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')
        
        if new_password != confirm_password:
            return jsonify({'success': False, 'message': '两次密码输入不一致'})
        
        if len(new_password) < 6:
            return jsonify({'success': False, 'message': '密码至少6位'})
        
        user = User.query.filter_by(username=session.get('username')).first()
        if not user or not verify_password(old_password, user.password_hash):
            return jsonify({'success': False, 'message': '原密码错误'})
        
        user.password_hash = hash_password(new_password)
        db.session.commit()
        return jsonify({'success': True, 'message': '密码修改成功'})
    
    return render_template('change_password.html')

@app.route('/user/<int:user_id>/role', methods=['POST'])
@login_required
def update_user_role(user_id):
    """更新用户角色（仅管理员）"""
    if session.get('role') != 'admin':
        return jsonify({'success': False, 'message': '无权限'})
    
    user = User.query.get_or_404(user_id)
    role = request.form.get('role', 'user')
    
    user.role = role
    db.session.commit()
    return jsonify({'success': True, 'message': '权限已更新'})

@app.route('/user/add', methods=['POST'])
@login_required
def add_user():
    """添加用户"""
    if session.get('role') != 'admin':
        return '无权限', 403
    username = request.form.get('username')
    password = request.form.get('password')
    role = request.form.get('role', 'user')
    
    if User.query.filter_by(username=username).first():
        return jsonify({'success': False, 'message': '用户名已存在'})
    
    user = User(username=username, password_hash=hash_password(password), role=role)
    db.session.add(user)
    db.session.commit()
    return redirect(url_for('users'))

@app.route('/user/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_user(user_id):
    """编辑用户"""
    if session.get('role') != 'admin':
        return '无权限', 403
    
    user = User.query.get_or_404(user_id)
    
    if request.method == 'POST':
        new_username = request.form.get('username')
        
        # 检查新用户名是否被占用（排除自己）
        if new_username != user.username and User.query.filter_by(username=new_username).first():
            return jsonify({'success': False, 'message': '用户名已存在'})
        
        user.username = new_username
        db.session.commit()
        return redirect(url_for('users'))
    
    return render_template('user_edit.html', user=user)

@app.route('/user/<int:user_id>/reset-password', methods=['GET', 'POST'])
@login_required
def reset_user_password(user_id):
    """重置用户密码（仅管理员）"""
    if session.get('role') != 'admin':
        return '无权限', 403
    
    user = User.query.get_or_404(user_id)
    
    if request.method == 'POST':
        new_password = request.form.get('password')
        if len(new_password) < 6:
            return jsonify({'success': False, 'message': '密码至少6位'})
        
        user.password_hash = hash_password(new_password)
        db.session.commit()
        return jsonify({'success': True, 'message': f'密码已重置为: {new_password}'})
    
    return render_template('user_reset_password.html', user=user)

@app.route('/user/<int:user_id>/delete', methods=['POST'])
@login_required
def delete_user(user_id):
    """删除用户"""
    if session.get('role') != 'admin':
        return '无权限', 403
    
    user = User.query.get_or_404(user_id)
    
    # 不允许删除自己
    if user.username == session.get('username'):
        return jsonify({'success': False, 'message': '不能删除自己'})
    
    # 不允许删除最后一个管理员
    if user.role == 'admin' and User.query.filter_by(role='admin').count() <= 1:
        return jsonify({'success': False, 'message': '不能删除最后一个管理员'})
    
    db.session.delete(user)
    db.session.commit()
    return redirect(url_for('users'))

# ============ 初始化 ============

def init_db():
    """初始化数据库"""
    with app.app_context():
        db.create_all()
        
        # 创建管理员账号
        if not User.query.filter_by(username='admin').first():
            admin = User(username='admin', password_hash=hash_password('admin123'), role='admin')
            db.session.add(admin)
        
        # 创建默认分类
        if Category.query.count() == 0:
            default_categories = ['办公用品', '电子设备', '工具', '耗材', '其他']
            for name in default_categories:
                db.session.add(Category(name=name))
        
        # 创建默认系统设置
        if SystemSettings.get('system_name') == '':
            SystemSettings.set('system_name', '物资管理系统')
        
        # 创建默认角色
        if SystemSettings.get('roles') == '':
            import json
            SystemSettings.set('roles', json.dumps([
                {'name': 'admin', 'is_admin': True, 'user_count': 0},
                {'name': 'user', 'is_admin': False, 'user_count': 0}
            ]))
            # 创建默认权限
            SystemSettings.set('permissions', json.dumps({
                'admin': {k: True for k in ['stock_in', 'stock_out', 'stock_check', 'products_view', 'products_edit', 'products_delete', 'categories_edit', 'users_view', 'users_edit', 'settings_view', 'settings_edit', 'reports_view']},
                'user': {'stock_in': True, 'stock_out': True, 'stock_check': True, 'products_view': True, 'products_edit': False, 'products_delete': False, 'categories_edit': False, 'users_view': False, 'users_edit': False, 'settings_view': False, 'settings_edit': False, 'reports_view': True}
            }))
        
        # 创建默认项目
        if Project.query.count() == 0:
            default_project = Project(
                name='默认项目',
                code='DEFAULT',
                description='系统默认项目'
            )
            db.session.add(default_project)
        
        # 创建 user_project_roles 表
        from sqlalchemy import text
        try:
            db.session.execute(text("""
                CREATE TABLE IF NOT EXISTS user_project_roles (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    project_id INTEGER NOT NULL,
                    username VARCHAR(50) NOT NULL,
                    role VARCHAR(50) DEFAULT 'viewer',
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (project_id) REFERENCES project(id),
                    UNIQUE(project_id, username)
                )
            """))
        except Exception:
            pass
        
        db.session.commit()
        print('数据库初始化完成')

if __name__ == '__main__':
    init_db()
    app.run(host='0.0.0.0', port=5000, debug=True)
